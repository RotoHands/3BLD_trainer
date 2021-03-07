import time
import keyboard
import pyperclip
import clipboard
import openpyxl
def removeSlesh(alg):
    if (alg[0] == '[' and alg[len(alg) - 1] == ']' and alg.count("[" )> 1):
        tempAlg = alg[:-1]
        newAlg = tempAlg[1:]
    else:
        newAlg = alg

    if(':' in newAlg):
        finalalg = newAlg.replace(":","")
    else:
        finalalg = newAlg
    return finalalg

def reverseSingal(letter):
    new=""
    length = len(letter)
    if (length == 1):
        return letter + "\'"
    if (length == 2):
        if (letter[1] == "2"):
            return letter
        elif (letter[1] == "\'"):
            return letter[0]
        else:
            return letter[0] + "\'" + letter[1] + "\'"
    if (length == 3):
        if(letter[2] == "\'"):
            if(letter[1] == "2"):
                return letter[0] + letter [1]
            else:
                return letter[0] + "\'" + letter[1]
        else:
            return letter[0] + letter[2] + "\'"
    else:
        return letter[0] + letter[2]

def makeAlg(All):
    A = All[0].split()
    B = All[1].split()
    C = All[2].split()
    Arev = []
    Brev = []
    Crev = []
    for x in A:
        Arev.append(reverseSingal(x))
    for x in B:
        Brev.append(reverseSingal(x))
    for x in C:
        Crev.append(reverseSingal(x))
    Arev.reverse()
    Brev.reverse()
    Crev.reverse()

    alg = C + A + B + Arev +  Brev + Crev

    return alg

def extendAlg(currentAlg):
    C=""
    if (currentAlg.find("[") != 0):
        C = currentAlg[:currentAlg.find("[")].strip()
        tempAlg = currentAlg[currentAlg.find("["):]
    else:
        tempAlg = currentAlg
    A = tempAlg[1:tempAlg.find(",")].strip()
    B = tempAlg[tempAlg.find(",") + 1:tempAlg.find("]")].strip()

    All =[A,B,C]
    return All

def cancel (f,s):
    if(len(f) == 1):
        score1 = 1
    else:
        score1 = [-1,2][f[1] == "2"]
    if (len(s) == 1):
        score2 = 1
    else:
        score2 = [-1, 2][s[1] == "2"]

    sum = score1 + score2
    if (sum == 0 or sum == 4):
        return ""
    if(sum == 1):
        return f[0]
    if (sum == 2 or sum == -2):
        return f[0] + "2"
    if(sum == -1 or sum == 3):
        return f[0] + "\'"

def reverseAlgCor(alg):
    reversealg =""
    algsplit = alg.split()
    algsplit.reverse()
    for x in algsplit:
        reversealg = reversealg + reverseSingal(x) + " "
    return reversealg

def cancelAlg(alg):
    for i in range (0,len(alg)-1):

        if(alg[i] != ""  and alg[i+1]!=""):
            if (alg[i][0] == alg[i+1][0] ):
                res = cancel(alg[i],alg[i+1])
                if (len(alg[i+1]) == 1):
                    alg[i+1] = ""


                else:
                    if(len(alg[i+1]) == 2):
                        if(alg[i+1][1] == "\'" or alg[i+1][1] == "2"):
                            alg[i + 1] =""

                        else:
                           alg[i + 1] = alg[i+1][1:]
                    else:
                        if (alg[i + 1][1] == "\'" or alg[i + 1][1] == "2"):
                            alg[i + 1] = alg[i + 1][2:]
                        else:
                            alg[i + 1] = alg[i + 1][1:]
                alg[i] = res



    return alg

def algmakerCor2():
    temp = ""
    final = ""

    clip = str(clipboard.paste())
    if (clip.find('w') != -1):
        wide_move = clip[clip.index('w')-1]
        alg = clip.replace(wide_move + 'w', wide_move.lower())
    else:
        alg = clip


    if (alg != "None"):
        if (alg.find("[") != -1):
            temp1 = removeSlesh(alg)
            newalg = makeAlg(extendAlg(temp1))
            temp = cancelAlg(newalg)
            final = ""
            for x in temp:
                final = final + x + " "

            final = final.replace("  ", " ")
            final = final.strip()

            print("normal: \n",final)
            print("")
            pyperclip.copy(final)
            print("reverse: \n", reverseAlgCor(final))

    return final

def transposeAndReverseTable(sheet_name):
    file = openpyxl.load_workbook(r"C:\Users\rotem\PycharmProjects\pythonProject\BLD\ROTO 3bld Algs.xlsx")
    sheet = file[sheet_name]
    for i in range (2,23):
        for j in range(2,i):

            currentAlg = sheet.cell(j,i).value
            if (currentAlg != None):
                reverseAlg = reverseAlgCor(currentAlg)
                sheet.cell(i,j).value = reverseAlg
                print(currentAlg)
    file.save(r"C:\Users\rotem\PycharmProjects\pythonProject\BLD\ROTO 3bld Algs.xlsx")

def f1():
    pyperclip.copy(reverseAlgCor(pyperclip.paste()))
keyboard.add_hotkey("ctrl+c", algmakerCor2())
keyboard.add_hotkey("ctrl+v", f1())
transposeAndReverseTable("edges")
"""
while(True):

    if(keyboard.is_pressed('ctrl+c')):
        time.sleep(0.2)
        alg = algmakerCor2()
        time.sleep(0.2)
        while(keyboard.is_pressed("ctrl+v") == False and keyboard.is_pressed("esc") == False):
            pass
        time.sleep(0.2)
        pyperclip.copy(reverseAlgCor(alg))
"""