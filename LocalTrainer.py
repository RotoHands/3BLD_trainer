import random
import time
import keyboard
import setting
from openpyxl import load_workbook
# coding=utf8
import openpyxl
import os

def waitS():

    space = False
    f = False
    while(space == False and f == False):
        space = keyboard.is_pressed('ctrl')
        f = keyboard.is_pressed('alt')
    time.sleep(0.2)
    if(space == True):
        return "next"
    if(f == True):
        return "forgot"
    else:
        return "error"

def findWord(LP,type):
    LP1=""
    if(len(LP) == 4):
        if(LP[1] == 'צ'):
            LP1 = '1' + '2'
        else:
            LP1 = '2' + '1'
    elif (len(LP) == 3):
        if(LP[1] == 'צ'):
            LP1 =  LP[0] + '1'
        elif(LP[1] == 'ג'):
            LP1 = LP[0] + '2'
        elif(LP[0] == 'צ'):
            LP1 = '1'+ LP[2]
        else:
            LP1 = '2'+ LP[2]
    else:
        LP1 = LP[0]+LP[1]
    LProw = setting.letters.index(LP1[0])*4+2
    LPcol = setting.letters.index(LP1[1])*3+int(type)

    cell = str(setting.sheet.cell(LProw, LPcol).value)
    wordsList=[]
    i=0
    while (i<=2):
        if (cell != "None"):
            wordsList.append(cell)
        LProw+=1
        cell = str(setting.sheet.cell(LProw, LPcol).value)
        i+=1

    return wordsList

def randomize():
    flag = False
    letters = ['א', 'ב', 'ג', 'ד', 'ה', 'ו', 'ז', 'ח', 'ט', 'י', 'כ', 'ל', 'מ', 'נ', 'ס', 'ע', 'פ', 'צ', 'ק', 'ר', 'ש','ת', 'צ\'', 'ג\'']
    pair =""
    while (flag == False):
        randomRow = random.randint(0, 23)
        randomCol = random.randint(0, 23)
        if(randomRow != randomCol):
            flag = True
            pair = letters[randomRow] + letters[randomCol]
    return pair

def checkWord(word):
    if word[0] in setting.letters and word[1] in setting.letters and (word[2] == "1" or word[2] == "2" or word[2] == "3"):
        return True
    return False

def printLetterPair2(pair):
    for i in range (1,577):

        if (setting.quizlet.cell(i,1).value == pair):
            print(setting.quizlet.cell(i,2).value)


def printLetterPair(pair):

    print("אדם: ", *findWord(pair, 1), sep=', ')
    print("פועל: ", *findWord(pair, 2), sep=', ')
    print("אובייקט: ", *findWord(pair, 3), sep=', ')

def writeLetterPair ():
    wb2 = load_workbook('LetterPairs.xlsx', data_only=True)
    quizlet = wb2["quizlet"]
    letters = ['א', 'ב', 'ג', 'ד', 'ה', 'ו', 'ז', 'ח', 'ט', 'י', 'כ', 'ל', 'מ', 'נ', 'ס', 'ע', 'פ', 'צ', 'ק', 'ר', 'ש',
               'ת', 'צ\'', 'ג\'']
    m = 1
    sent=""
    print("start")
    for i in range (0, len(setting.letters)):
        for j in range (0, len(setting.letters)):
            pair = setting.letters[i] + setting.letters [j]
            print(pair)
            for n in range (1,4):
                partSent = findWord(pair, n)
                for x in partSent:
                    sent += x
                sent += " "
            print(sent)
            quizlet.cell(m,2).value = sent
            quizlet.cell(m, 1).value = letters[i] + letters[j]
            m+=1
            sent = ""

    wb2.save('LetterPairs.xlsx')
def countAll():
    c=0
    for i in range (1,96):
        for j in range (1,78):
            cell = str(setting.sheet.cell(i, j).value)
            wordsList = []
            if cell != "None" :
                c+=1
                print(c)
                print(cell)
    print ("all pairs number is : ",c)
def trainMemory():
   c=0
   while(keyboard.is_pressed('q') == False):
       pair = randomize()
       x = "." + str(c)+" "+ "זוג האותיות הוא: " + pair
       print(x)
       waitS()
       printLetterPair(pair)
       c+=1
       print("")
def countD():
    wb = load_workbook('LetterPairs.xlsx', data_only=True)
    words = wb["quizlet"]
    c=0
    d=0
    for i in range (1,577):
        for j in range (0,3):
            if(words.cell(i,j+10).value != "no"):
                c+=1

    return c
def resetMemory():
    wb = load_workbook('LetterPairs.xlsx', data_only=True)
    words = wb["quizlet"]

    for i in range (0,3):
        for j in range (1,576):
            if(words.cell(j,i+10).value == "yes"):
                words.cell(j, i + 10).value = "no"
    wb.save('LetterPairs.xlsx')
def fixTable():
    wb = load_workbook('LetterPairs.xlsx', data_only=True)
    words = wb["quizlet"]
    for j in range (1,576):
        if (words.cell(j,12).value == "none"):
            words.cell(j, 6).value == words.cell(j, 4).value
            words.cell(j, 12).value = "no"
            words.cell(j, 10).value = "none"
    wb.save('LetterPairs.xlsx')

def fixTable2():
    wb = load_workbook('LetterPairs.xlsx', data_only=True)
    words = wb["quizlet"]
    for j in range (1,576):
        if (words.cell(j,10).value == "none"):
            print("here 1")
            words.cell(j, 6).value = words.cell(j, 4).value
            words.cell(j, 4).value = ""
    wb.save('LetterPairs.xlsx')
def trainMemory2():
    wb = load_workbook('LetterPairs.xlsx', data_only=True)
    words = wb["quizlet"]
    finish = 1671
    count = 0

    flag = False
    results = []
    while (keyboard.is_pressed('q') == False):

        currentSen = []
        currentSenNum = []
        ran  =  random.randint(1,576)
        for i in range (0,3):
            if(flag == True):
                break
            while(words.cell(ran,i+10).value != "no" and flag == False):
                if(count == finish):
                    flag = True
                ran = random.randint(1,576)

            if (flag == False):
                currentSen.append(words.cell(ran,1).value)
                currentSenNum.append(ran)
                words.cell(ran,i+10).value = "yes"
                ran = random.randint(1, 576)
                count+=1
            else:
                print("here 1")
        print(*currentSen, sep=" ")
        startTime = time.time()
        key = waitS()
        if (key == "next"):
            res = time.time() - startTime
            results.append([currentSenNum,res])
        if (key == "forgot"):
            sen = ""
            j=0
            for word in currentSenNum:
                sen += words.cell(word,4+j).value+" "
                j+=1
            print(sen+"\n")

    countTrain=0
    for result in results:
        print(result)
        for pair in result[0]:
            countTrain +=1
            index = pair
            print("index is", index)
            avr = ((words.cell(index,15).value * words.cell(index,16).value) + result[1])/(words.cell(index,16).value+1)
            words.cell(index, 15).value = avr
            words.cell(index, 16).value +=1
    print("trained: ", countTrain/3)
    wb.save('LetterPairs.xlsx')

def findIndex(pair):
    wb = load_workbook('LetterPairs.xlsx', data_only=True)
    words = wb["quizlet"]
    print("pair is :", pair)
    for i in range (1,576):
        if (words.cell(i,1).value == pair):
            return i

    print("error")

def addSpace():
    wb = load_workbook('LetterPairs.xlsx', data_only=True)
    words = wb["quizlet"]

    for i in range(1, 577):
        for j in range (0,3):
            if(words.cell(i,j+4).value!= None):
                find = words.cell(i,j+4).value.find("(")
                print("hello")
                if( find != -1 and words.cell(i,j+4).value[find-1] != " " ):
                    print("addspaces")
                    words.cell(i, j + 4).value = words.cell(i,j+4).value[:find] + " " + words.cell(i,j+4).value[find:]
    wb.save('LetterPairs.xlsx')

def fixLetterPair():
    wb = load_workbook('LetterPairs.xlsx', data_only=True)
    words = wb["quizlet"]

    for i in range(1, 577):
        newSen = []
        z=0
        for j in range (0,10):
            print(words.cell(i,j+4).value )
            if(words.cell(i,j+4).value != None):
                print("here 1")
                print(words.cell(i, j + 4).value)
                if (words.cell(i,j+4).value.find("(") == 0):
                    print("new Sen is: ", *newSen, sep = ", ")
                    print(j-1)
                    newSen[j-1-z] = newSen[j-1-z] + words.cell(i,j+4).value
                    z+=1

                else:
                    newSen.append(words.cell(i,j+4).value)
        m = 15
        for word in newSen:
            words.cell(i,m).value = word
            m+=1
    wb.save('LetterPairs.xlsx')



def genTrainMemory():
    wb = load_workbook('LetterPairs.xlsx', data_only=True)
    words = wb["quizlet"]

    for i in range(1,577):
        currentLetterPairs = words.cell(i,2).value.split()
        newSen =[]
        startSog = False
        endSog = False
        sograim = ""
        for word in currentLetterPairs:

            if (word.find("(") != -1):
                startSog = True
            if(word.find(")") != -1):
                endSog = True
            if (startSog == True):
                sograim +=word +" "
            else:
                newSen.append(word)
            if(endSog == True):
                newSen.append(sograim)
                startSog = False
                endSog = False
                sograim=""
        j=0
        for word in newSen:
            words.cell(i,4+j).value = word
            j+=1
        newSen = []

    wb.save('LetterPairs.xlsx')

def trainMemoryWithAlgs(pair):
   printLetterPair(pair)
def resetTimes():
    wb = load_workbook('LetterPairs.xlsx', data_only=True)
    words = wb["quizlet"]

    for i in range (1,576):
        words.cell(i,15).value = 0
        words.cell(i, 16).value = 0
    wb.save('LetterPairs.xlsx')

def genList():
    wb = load_workbook('LetterPairs.xlsx', data_only=True)
    words = wb["quizlet"]
    list = "["
    for i in range(1,576):
        list+="\"" + words.cell(i,1).value +"\""+ ","

    list+=" end" + "]"
    print(list)


def printWord(word):
        if (checkWord(word)):
            wordsList = findWord(word[:2], word[2])
            print(*wordsList, sep='\n')
        else:
            print("קלט שגוי, הכנס מילה")


def genRandomList():
    wb = load_workbook('LetterPairs.xlsx', data_only=True)
    words = wb["quizlet"]
    finish = 1671
    count = 0

    flag = False
    all = []
    results = []
    l = 0
    while (keyboard.is_pressed('q') == False):

        currentSen = []
        term = []
        defin = []
        currentSenNum = []
        ran  =  random.randint(1,576)

        l+=1
        print("count is :", l)
        for i in range (0,3):
            if(flag == True):
                break
            while(words.cell(ran,i+10).value != "no" and flag == False):
                if(count == finish):
                    flag = True
                ran = random.randint(1,576)

            if (flag == False):
                currentSen.append(words.cell(ran,1).value)
                currentSenNum.append(ran)
                #words.cell(ran,i+10).value = "yes"
                ran = random.randint(1, 576)
                count+=1
            else:
                print("here 1")
        sen =""
        for word in currentSen:
            sen += word + " "
        t=0
        definition=""
        for index in currentSenNum:
            print("index is:", index)
            definition += words.cell(index,4+t).value + " "
            t+=1
        defin.append(definition)
        term.append(sen)
        all.append([term,defin])
    k=1
    for sen in all:
        print("sen is: ", sen)

        words.cell(k,20).value = sen[0][0]
        words.cell(k, 21).value = sen[1][0]
        k+=1
    print("here 1")
    wb.save('LetterPairs.xlsx')

