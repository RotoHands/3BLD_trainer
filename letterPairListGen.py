

import openpyxl
from openpyxl import  load_workbook

wb = load_workbook("LetterPairs.xlsx")
ws = wb["quizlet"]
current = ""
write2txt = open("letter.txt", "w+")
for i in range (1,577):
    current = ""
    for j in range(2,4):
        if (ws.cell(i,j).value != None):
            current+=ws.cell(i,j).value +" "
    write2txt.write(current[:len(current)-1])