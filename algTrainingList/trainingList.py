#-*- coding: utf-8 -*-
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def genAlgListEd():
    letterTrainEd = open("trainingSetEdges.txt", "r", encoding="utf8")
    i = 1
    for line in letterTrainEd:

        edTrain.cell(i, 1).value = lettersEdges[lettersEdgesTrain.index(line[0])] + lettersEdges[
            lettersEdgesTrain.index(line[1])]

        edTrain.cell(i, 2).value = edAlgs.cell(lettersEdgesTrain.index(line[0]) + 2,lettersEdgesTrain.index(line[1]) + 2).value
        i += 1
    wb.save(r"C:\Users\rotem\PycharmProjects\pythonProject\BLD\ROTO 3bld Algs.xlsx")


def genAlgListCor():
    letterTrainCor = open("trainingSetCorners.txt", "r", encoding="utf8")
    i = 1
    for line in letterTrainCor:

        corTrain.cell(i, 1).value = lettersCorners[lettersCornersTrain.index(line[0])] + lettersCorners[
            lettersCornersTrain.index(line[1])]

        corTrain.cell(i, 2).value = corAlgs.cell(lettersCornersTrain.index(line[1]) + 2,
                                                 lettersCornersTrain.index(line[0]) + 2).value
        i += 1
    wb.save(r"C:\Users\rotem\PycharmProjects\pythonProject\BLD\ROTO 3bld Algs.xlsx")
def resetEdFinal():
    allEdAlgs = open("allEdAlgs.txt", "w+")
    for i in range(2, 24):
        for j in range(2, 24):
            cell = edAlgs.cell(i, j)
            if (cell.fill.fgColor.index != "FFD0E0E3"):
                if (i==22):
                    if(j==23):
                        allEdAlgs.write("1" + "2" + "\n")
                    else:
                        allEdAlgs.write("1" + lettersEdges[j - 2] + "\n")
                elif (i == 23):
                    if (j == 22):
                        allEdAlgs.write("2" + "1" + "\n")
                    else:
                        allEdAlgs.write("2" + lettersEdges[j - 2] + "\n")
                elif (j==22):
                    allEdAlgs.write(lettersEdges[i - 2] + "1" + "\n")
                elif(j==23):
                    allEdAlgs.write( lettersEdges[i - 2] + "2" + "\n")
                else:
                    allEdAlgs.write(lettersEdges[i - 2] + lettersEdges[j - 2] + "\n")
    allEdAlgs.close()
    wb.save(r"C:\Users\rotem\PycharmProjects\pythonProject\BLD\ROTO 3bld Algs.xlsx")

wb = load_workbook(r"C:\Users\rotem\PycharmProjects\pythonProject\BLD\ROTO 3bld Algs.xlsx", data_only=True)

edAlgs = wb["edges"]
corAlgs = wb["corners"]
corD = wb["corners - drill"]
edD = wb["edges - drill"]
edTrain = wb["edgesTraining"]
corTrain = wb["cornersTraining" ]
lettersCorners = ['א', 'ב', 'ד', 'ה', 'ו', 'ז', 'ח', 'ט', 'כ', 'ל', 'נ', 'ס', 'ע', 'פ', 'צ', 'ק', 'ר', 'ש', 'ת',
                  'צ\'', 'ג\'']
lettersEdges = ['א', 'ב', 'ד', 'ה', 'ו', 'ז', 'ח', 'י', 'כ', 'ל', 'מ', 'נ', 'ס', 'ע', 'פ', 'צ', 'ק', 'ר', 'ש', 'ת',
                'צ\'', 'ג\'']
letters = ['א', 'ב', 'ג', 'ד', 'ה', 'ו', 'ז', 'ח', 'ט', 'י', 'כ', 'ל', 'מ', 'נ', 'ס', 'ע', 'פ', 'צ', 'ק', 'ר', 'ש',
           'ת', '1', '2']

lettersCornersTrain = ['א', 'ב', 'ד', 'ה', 'ו', 'ז', 'ח', 'ט', 'כ', 'ל', 'נ', 'ס', 'ע', 'פ', 'צ', 'ק', 'ר', 'ש', 'ת',
                  '1', '2']
lettersEdgesTrain = ['א', 'ב', 'ד', 'ה', 'ו', 'ז', 'ח', 'י', 'כ', 'ל', 'מ', 'נ', 'ס', 'ע', 'פ', 'צ', 'ק', 'ר', 'ש', 'ת',
                '1', '2']

def setEdgesForTraining():

    for i in range (2,24):
        for j in range(2,24):
            edD.cell(i,j).fill = PatternFill(fgColor="FFD0E0E3", fill_type = "solid")

    traingingSetEd = open(r"C:\Users\rotem\PycharmProjects\pythonProject\BLD\algTrainingList\trainingSetEdges.txt", "r", encoding="utf8")
    for line in traingingSetEd:
        edD.cell(lettersEdgesTrain.index(line[1])+2,lettersEdgesTrain.index(line[0])+2).fill = PatternFill(fgColor="00FFFFFF", fill_type="solid")
    wb.save(r"C:\Users\rotem\PycharmProjects\pythonProject\BLD\ROTO 3bld Algs.xlsx")


def setCorForTraining():
    for i in range(2, 23):
        for j in range(2, 23):
            corD.cell(i, j).fill = PatternFill(fgColor="FFD0E0E3", fill_type="solid")

    traingingSetCor = open(r"C:\Users\rotem\PycharmProjects\pythonProject\BLD\algTrainingList\trainingSetCorners.txt", "r",
                          encoding="utf8")
    for line in traingingSetCor:
        corD.cell(lettersCornersTrain.index(line[0]) + 2,lettersCornersTrain.index(line[1]) + 2).fill = PatternFill(
            fgColor="00FFFFFF", fill_type="solid")
    wb.save(r"C:\Users\rotem\PycharmProjects\pythonProject\BLD\ROTO 3bld Algs.xlsx")

setCorForTraining()
setEdgesForTraining()