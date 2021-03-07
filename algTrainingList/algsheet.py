#-*- coding: utf-8 -*-
import openpyxl
from openpyxl import load_workbook
wb = load_workbook(r'C:\Users\rotem\PycharmProjects\pythonProject\BLD\ROTO 3bld Algs.xlsx', data_only=True)
corAlgs = wb["corners"]
edAlgs = wb["edges"]
corAll = wb["allCorList"]
edAll = wb["allEdList"]
edprint = wb["edges_drill_print"]
cor_print = wb["cor_train_print"]
lettersCorners = ['א', 'ב', 'ד', 'ה', 'ו', 'ז', 'ח', 'ט', 'כ', 'ל', 'נ', 'ס', 'ע', 'פ', 'צ', 'ק', 'ר', 'ש', 'ת',
                  '1', '2']
lettersEdges = ['א', 'ב', 'ד', 'ה', 'ו', 'ז', 'ח', 'י', 'כ', 'ל', 'מ', 'נ', 'ס', 'ע', 'פ', 'צ', 'ק', 'ר', 'ש', 'ת',
                '1', '2']

def edges():
    ed_drill = open("edges_to_learn_final.txt", "r",encoding="utf-8")
    ed_drill_list = ed_drill.readlines()
    count  = 1
    for alg in ed_drill_list:
        edprint.cell(count , 1 ).value = alg
        edprint.cell(count , 2).value = edAlgs.cell(lettersEdges.index(alg[1])+2, lettersEdges.index(alg[0])+2).value
        count += 1
    wb.save(r'C:\Users\rotem\PycharmProjects\pythonProject\BLD\ROTO 3bld Algs.xlsx')

def corners():
    cor_drill = open("Corners_To_Learn.txt", "r",encoding="utf-8")
    cor_drill_list = cor_drill.readlines()
    count  = 1
    for alg in cor_drill_list:
        cor_print.cell(count , 1 ).value = alg
        cor_print.cell(count , 2).value = corAlgs.cell(lettersCorners.index(alg[1])+2, lettersCorners.index(alg[0])+2).value
        count += 1
    wb.save(r'C:\Users\rotem\PycharmProjects\pythonProject\BLD\ROTO 3bld Algs.xlsx')

corners()