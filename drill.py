import openpyxl
import random
from shutil import copyfile

import time
from LocalTrainer import printWord
import keyboard
from LocalTrainer import waitS
import setting
from algClass import Alg
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager





"""
def genScra():
    browser = webdriver.Chrome(ChromeDriverManager().install())
    browser.get("https://cstimer.net")
    while(True):
        a=input("press enter to continue")
        nex = browser.find_elements_by_class_name("click")[61]
        print(nex)
        scra = browser.find_element_by_id("scrambleTxt")
        file = open("scra.txt", "a")
        for i in range (3000):
            print((i))
            nex.click()
            if (scra.text!="Scrambling..."):
                file.write(scra.text+"\n")

        file.close()
"""

def genAlg(i,piece):

    if (piece == "corners"):
        a = [[setting.CorTrain[i%setting.numlineC][0],setting.CorTrain[i%setting.numlineC][1]],setting.corAlgs.cell(setting.CorTrain[i%setting.numlineC][0],setting.CorTrain[i%setting.numlineC][1]).value,setting.lettersCorners[setting.CorTrain[i%setting.numlineC][1]-2]+setting.lettersCorners[setting.CorTrain[i%setting.numlineC][0]-2]]
        return a

    elif (piece == "edges"):
        b = [[setting.EdTrain[i%setting.numlineE][0],setting.EdTrain[i%setting.numlineE][1]],setting.edAlgs.cell(setting.EdTrain[i%setting.numlineE][0],setting.EdTrain[i%setting.numlineE][1]).value, setting.lettersEdges[setting.EdTrain[i%setting.numlineE][1]-2]+setting.lettersEdges[setting.EdTrain[i%setting.numlineE][0]-2]]
        return b
    else:
        print("error 2")

def insertResultCor(index, time):
    col = index[1]
    row = index[0]
	
    currentAvg = setting.corD.cell(row,col).value
    currentTimes = setting.corD.cell(row + 27,col).value
    newAvg = float(((currentAvg*currentTimes) + time )/(currentTimes + 1))
    currentTimes +=1
    setting.corD.cell(row, col).value = newAvg
    setting.corD.cell(row + 27, col).value = currentTimes

def insertResultEdge(index, time):
    col = index[1]
    row = index[0]
	
    currentAvg = setting.edD.cell(row,col).value
    currentTimes = setting.edD.cell(row + 27,col).value
    newAvg = float(((currentAvg*currentTimes) + time )/(currentTimes + 1))
    currentTimes +=1
    setting.edD.cell(row, col).value = newAvg
    setting.edD.cell(row + 27, col).value = currentTimes

def bldAttempt():
    print("Start attempt!")
    word =""
    while(word != "סיום"):
        word = input()
        printWord(word)
        print()

def saveResults(results,piece):
    pass

    #if(piece == "corners"):
        #for x in results:
            #insertResultCor(x[0],x[1])
    #elif (piece == "edges"):
        #for x in results:
            #insertResultEdge(x[0], x[1])

    #setting.wb.save("ROTO 3bld Algs.xlsx")


def saveResNew(piece):
    wb = openpyxl.load_workbook("ROTO 3bld Algs.xlsx")
    wsC = wb["corners - drill"]
    wsE = wb["edges - drill"]
    wsE.cell(1,1).value = wsE.cell(1,1).value #make sure xlsx is writable
    if(piece == "corners"):
        file = open('timesCor.txt', 'r')
    elif(piece == "edges"):
        file = open('timesEd.txt', 'r')
    else:
        print("erorr!!")

    copyfile("timesCor.txt", "timesCorBackUp.txt")
    copyfile("timesEd.txt", "timesCorBackUp.txt")

    allResults = file.readline().split(';')
    print("allres", *allResults, sep="\n")

    if (piece == "corners"):
        print(len(allResults))
        for res in allResults:
            if(len(res) ==0):
                break
            split = res.split(',')
            row = int(split[0])
            col = int(split[1])
            time = float(split[2])
            currentAvg = wsC.cell(row, col).value
            currentTimes = wsC.cell(row + 27, col).value
            newAvg = float(((float(currentAvg) *float(currentTimes)) + time) / (float(currentTimes) + 1))
            currentTimes += 1
            print ("row ", row, " ,col ", col," ,time", time, ", current avg", currentAvg," ,new avg", newAvg)

            wsC.cell(row, col).value = newAvg
            wsC.cell(row + 27, col).value = currentTimes
            open('timesCor.txt','w').close()
    else:
        print(len(allResults))
        for res in allResults:
            if(len(res) ==0):
                break
            split = res.split(',')
            row = int(split[0])
            col = int(split[1])
            time = float(split[2])

            currentAvg = wsE.cell(row, col).value
            currentTimes = wsE.cell(row + 27, col).value
            newAvg = float(((float(currentAvg) *float(currentTimes)) + time) / (float(currentTimes) + 1))
            currentTimes += 1
            wsE.cell(row, col).value = newAvg
            wsE.cell(row + 27, col).value = currentTimes
            open('timesEd.txt', 'w').close()
    wb.save("ROTO 3bld Algs.xlsx")
saveResNew("corners")
saveResNew("edges")






