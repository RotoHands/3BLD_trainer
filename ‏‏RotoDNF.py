from bleak import BleakClient
import asyncio
import time
from openpyxl import load_workbook
import openpyxl
import permutation
import copy
import keyboard
import webbrowser

def scrambleGen():
    scramblesTxt  = open ("scrambles.txt", "r")
    content = scramblesTxt.readlines()
    data =[]
    for x in content:
        i =x.find(",")
        if(i!= -1):
            new0 = x[i+2:]
        else:
            new0 = x
        wide1 = new0.find("w")
        if (wide1 != -1):
            lowerCase = chr(ord(new0[wide1-1]) + 32)
            new1 = new0.replace("w",lowerCase,1)
            new2 = new1[:wide1-1] + new1[wide1:]
            data.append(new2)
        else:
            data.append(new0)

    scramblesTxt.close()

def deleteWideMoves():
    scramblesTxt = open("scrambles.txt", "r")
    content = scramblesTxt.readlines()
    content = [x.strip() for x in content]
    data = []
    finalData = []
    for scramble in content:
        i=0
        lowerPlace = 0
        for letter in scramble:
            if(letter.islower() == True):
                lowerPlace = i
            i+=1
        if(lowerPlace!=0):
            data.append(scramble[:lowerPlace])
        else:
            data.append(scramble)

    print(*data,sep ="\n")
    scramblesTxt.close()

def checkEndScramble(moves):
    if (len(moves) < 4):
        return False
    size = len(moves)

    if (moves[size - 1] == "R" and moves[size - 2] == "R" and moves[size - 3] == "R" and moves[size - 4] == "R"):
        return True
    if (moves[size - 1] == "R" and moves[size - 2] == "R" and moves[size - 3] == "R2"):
        return True
    if (moves[size - 1] == "R" and moves[size - 2] == "R2" and moves[size - 3] == "R"):
        return True
    if (moves[size - 1] == "R2" and moves[size - 2] == "R" and moves[size - 3] == "R"):
        return True
    if (moves[size - 1] == "R2" and moves[size - 2] == "R2"):
        return True
    if (moves[size - 1] == "D" and moves[size - 2] == "D" and moves[size - 3] == "D" and moves[size - 4] == "D"):
        return True
    if (moves[size - 1] == "D" and moves[size - 2] == "D" and moves[size - 3] == "D2"):
        return True
    if (moves[size - 1] == "D" and moves[size - 2] == "D2" and moves[size - 3] == "D"):
        return True
    if (moves[size - 1] == "D2" and moves[size - 2] == "D" and moves[size - 3] == "D"):
        return True
    if (moves[size - 1] == "D2" and moves[size - 2] == "D2"):
        return  True
    return False

def checkEndSolve(Cube):
    moves = Cube.moves
    if (len(moves) < 4):
        return False
    size = len(moves)
    if (moves[size - 1] == "R" and moves[size - 2] == "R" and moves[size - 3] == "R" and moves[size - 4] == "R"):
        return True
    if (moves[size - 1] == "R" and moves[size - 2] == "R" and moves[size - 3] == "R2"):
        return True
    if (moves[size - 1] == "R" and moves[size - 2] == "R2" and moves[size - 3] == "R"):
        return True
    if (moves[size - 1] == "R2" and moves[size - 2] == "R" and moves[size - 3] == "R"):
        return True
    if (moves[size - 1] == "R2" and moves[size - 2] == "R2"):
        return True
    if (moves[size - 1] == "D" and moves[size - 2] == "D" and moves[size - 3] == "D" and moves[size - 4] == "D"):
        return True
    if (moves[size - 1] == "D" and moves[size - 2] == "D" and moves[size - 3] == "D2"):
        return True
    if (moves[size - 1] == "D" and moves[size - 2] == "D2" and moves[size - 3] == "D"):
        return True
    if (moves[size - 1] == "D2" and moves[size - 2] == "D" and moves[size - 3] == "D"):
        return True
    if (moves[size - 1] == "D2" and moves[size - 2] == "D2"):
        return True
        return  True
    if(Cube.alg.isSolved == True):
        print("solved!!")
        return True

    return False


def cleanRep (moves,Cube):
    print("moves are: ", *moves, sep=" ")
    sol = moves
    i=0
    while (i< len(sol) - 1):
        check = checkTwoSum(sol[i], sol[i+1])
        if(check[0] == "sumZero"):
            if((len(sol) - 2) == i):
                sol = sol[:i]
                Cube.timeMoves = Cube.timeMoves[:i]
            else:
                sol = sol[:i] + sol[i+2:]
                Cube.timeMoves = Cube.timeMoves[:i] + Cube.timeMoves[i+2:]
            i=0
        elif(check[0] == "noChange"):
            i+=1
        elif(check[1] == True):
            sol[i] = check[0]
            if ((len(sol) - 2) == i):#before last
                sol = sol[:i+1]
                Cube.timeMoves = Cube.timeMoves[:i+1]
            elif(i < len(sol) - 2):#regular
                sol = sol[:i+1] + sol[i+2:]
                Cube.timeMoves = Cube.timeMoves[:i+1] + Cube.timeMoves[i + 2:]
            else:#last
                pass
            i=0
        else:
            i+=1
    print ("after clean: ", *sol, sep=" ")
    return sol
class DNFanalyzer:
    normalDiff = []
    sliceDiff = []
    alg = Alg("")
    moves = []
    timeMoves = []
    scrambleToExe = ""
    scrambleApplied = ""
    firstMove = True
    firstSolveMove = True
    scramble = True
    memo = False
    solving = False
    startMemoTime = time.time()
    memoTime = time.time()
    startExeTime = time.time()
    exeTime = time.time()
    isSolvedEdges = False
    isSolvedCorners = False
    bestArrayCor = []
    bestArrayEdge = []
    bestCorOrder = 0
    bestEdgeOrder = 0
    startCorState = alg.corenerState
    startEdgeState = alg.edgeState
    mostPiecesSolved = 0
    moveNumber = 0
    solutionEdges = []
    solutionCorners = []
    Recon = ""
    rawSol = []
    success = False
def insertScramble(Cube):
    scramble =""
    for x in Cube.moves:
        scramble = scramble + x +" "
    Cube.scrambleApplied = scramble
    print("scrammble applied" ,scramble)

def executeScramble(Cube):
    movesToExecute = ""
    for move in Cube.moves:
        movesToExecute = movesToExecute + move + " "

    Cube.alg.movesToExecute = movesToExecute
    Cube.alg.executeAlg()
    Cube.alg.reverseSelf()
    Cube.alg.printState()
    Cube.startEdgeState = Cube.alg.edgeState
    Cube.startCorState = Cube.alg.corenerState

def checkSolEdges(Cube):

    currentCorState = Cube.alg.corenerState

    currentEdgeorder = Cube.alg.countSolveEdges()
    if (currentCorState == Cube.startCorState and currentEdgeorder > Cube.bestEdgeOrder):
        print("currentEdgeOrder: ", currentEdgeorder," bestEdgeOrder :", Cube.bestEdgeOrder )
        Cube.bestArrayEdge.append(Cube.moveNumber)
        Cube.bestEdgeOrder = currentEdgeorder
        Cube.mostPiecesSolved = Cube.moveNumber


def checkSolCoreners(Cube):
    currentEdgeState = Cube.alg.edgeState
    currentCororder = Cube.alg.countSolvedCor()


    if (currentEdgeState == Cube.startEdgeState and currentCororder > Cube.bestCorOrder):
        print("currentCorOrder: ", currentCororder, " bestCorOrder :", Cube.bestCorOrder)
        Cube.bestArrayCor.append(Cube.moveNumber)
        Cube.bestCorOrder = currentCororder
        Cube.mostPiecesSolved = Cube.moveNumber
    if(Cube.alg.countSolvedCor() == 8):
        Cube.isSolvedCorners = True

def isInArray(array,i):
    for j in array:
        if(j == i):
            return True
    return False

def insertSolveStats(Cube):
    wb = load_workbook('RotoDNFStats.xlsx', data_only=True)
    stats = wb["RotoStats"]
    i=2
    print("herre 3")
    while(stats.cell(i,2).value != "no"):
       i+=1
    stats.cell(i,2).value = "yes"
    stats.cell(i, 3).value = Cube.memoTime
    stats.cell(i, 4).value = Cube.exeTime

    if(Cube.success == True):
        stats.cell(i, 5).value = "yes"
    else:
        stats.cell(i,5).value = "no"
    print("herre 3")
    stats.cell(i, 6).value = Cube.Recon
    stats.cell(i, 7).value = Cube.Recon


    wb.save('RotoDNFStats.xlsx')
def getNewScramble():
    wb = load_workbook('RotoDNFStats.xlsx', data_only=True)
    stats = wb["RotoStats"]
    i = 2
    while (stats.cell(i, 2).value != "no"):
        i += 1

    wb.save('RotoDNFStats.xlsx')
    print(stats.cell(i,1).value)
    return stats.cell(i,1).value

def analyzeSolve(Cube):
    Cube.alg.reset()
    Cube.alg.movesToExecute = Cube.scrambleApplied
    Cube.alg.executeAlg()
    Cube.alg.reverseSelf()
    Cube.startCorState = Cube.alg.corenerState
    Cube.alg.printState() #good
    Cube.moveNumber = 0
    print("sol Edges1: ", *Cube.solutionEdges, sep=" ")
    Cube.alg.movesToExecute = []
    finalSolMoves = []
    for move in Cube.solutionEdges:
        finalSolMoves.append(move)
    for move in Cube.solutionCorners:
        finalSolMoves.append(move)

    print("sol Edges: ", *Cube.solutionEdges, sep =" ")
    print("sol Corners: ", *Cube.solutionCorners, sep = " ")
    print("sol raw : ", *Cube.rawSol, sep = " ")


    for move in Cube.solutionEdges:
        Cube.moveNumber+=1
        Cube.alg.movesToExecute = move
        Cube.alg.reverseSelf()
        Cube.alg.executeAlg()
        Cube.alg.reverseSelf()

        checkSolEdges(Cube)

    print("bestArrayEdge: ", *Cube.bestArrayEdge, sep=" ")
    Cube.mostPiecesSolved = Cube.bestArrayEdge[len(Cube.bestArrayEdge) - 1]
    print("slice Diff", *Cube.sliceDiff, sep = "\n")
    print("normal Diff", *Cube.normalDiff, sep = "\n")




    Cube.startEdgeState = Cube.alg.edgeState

    print("move number cor", Cube.moveNumber)
    print("sol corners: ", *Cube.solutionCorners, sep = " ")
    for move in Cube.solutionCorners:
        Cube.moveNumber+=1
        Cube.alg.movesToExecute = move
        Cube.alg.reverseSelf()
        Cube.alg.executeAlg()
        Cube.alg.reverseSelf()
        checkSolCoreners(Cube)



    if( len(Cube.bestArrayCor) > 0 ):

        #print("bestArrayCor: ", *Cube.bestArrayCor, sep = " ")
        #print("mostPic: ", Cube.bestArrayCor[len(Cube.bestArrayCor) - 1])
        Cube.mostPiecesSolved  = Cube.bestArrayCor[len(Cube.bestArrayCor)-1]
    mistake = 0
    if(Cube.isSolvedEdges == True):
        if(Cube.isSolvedCorners == True): #good solve
            Cube.success = True
            print("success!!")
        else:
            if (len(Cube.bestArrayCor) > 0):
                mistake = Cube.bestArrayCor[len(Cube.bestArrayCor) - 1]
            else:
                mistake = 0
    else:
        if (len(Cube.bestArrayEdge) > 0):
            mistake = Cube.bestArrayEdge[len(Cube.bestArrayEdge) - 1]
        else:
            mistake = 0



    print("best places: Cor", *Cube.bestArrayCor, sep = " ")
    print("best places: Edge", *Cube.bestArrayEdge, sep=" ")
    placeBeforeMistake = mistake
    print("heer 1")
    urlAlgCubing = "https://alg.cubing.net/?setup="
    for move in Cube.scrambleApplied.split():
        if(len(move) == 2):
            if(move[1] == "\'"):
                urlAlgCubing += move[0] +"-"
            else:
                urlAlgCubing += move[0] + "2"
        else:
            urlAlgCubing+= move
        urlAlgCubing+="_"

    urlAlgCubing+="&alg="
    if (mistake == 0):
        mistake = -1

    for i in range(0, len(finalSolMoves)):
        if (i != mistake):
            if (isInArray(Cube.bestArrayEdge, i) or isInArray(Cube.bestArrayCor, i)):
                urlAlgCubing += "%0A"
            if (len(finalSolMoves[i]) == 2):
                if (finalSolMoves[i][1] == "\'"):
                    urlAlgCubing += finalSolMoves[i][0] + "-"
                else:
                    urlAlgCubing += finalSolMoves[i][0] + "2"
            else:
                urlAlgCubing += finalSolMoves[i]

        else:
            urlAlgCubing += "%0A%2F%2F_mistake_starts_here%2F%2F%0A"
            if (len(finalSolMoves[i]) == 2):
                if (finalSolMoves[i][1] == "\'"):
                    urlAlgCubing += finalSolMoves[i][0] + "-"
                else:
                    urlAlgCubing += finalSolMoves[i][0] + "2"
            else:
                urlAlgCubing += finalSolMoves[i]
        urlAlgCubing += "_"
    print("heer 5")
    chrome_path = 'C:/Program Files (x86)/Google/Chrome/Application/chrome.exe %s'
    Cube.Recon = urlAlgCubing
    webbrowser.get(chrome_path).open(urlAlgCubing)


def isScrambleCorrct(Cube):#checks if the scramble is correct
    currentState = Alg("")
    for move in Cube.moves:
        currentState.movesToExecute = currentState.movesToExecute + move + " "
    currentState.executeAlg()
    currentState.reverseSelf()
    scramble = Alg("")
    scramble.movesToExecute = Cube.scrambleToExe
    scramble.executeAlg()
    scramble.reverseSelf()
    for i in range(1,25):
        if(scramble.edgeState(i) != currentState.edgeState(i) or scramble.corenerState(i) != currentState.corenerState(i)):
            return False
    return  True

def genUrlSol(Cube, moves, urlAlgCubing, mistake):#gens the link to algCubing.net of the solve
    finalSolMoves = moves
    for i in range(0, len(finalSolMoves)):
        if (i != mistake):
            if (isInArray(Cube.bestArrayEdge, i) or isInArray(Cube.bestArrayCor, i)):
                urlAlgCubing += "%0A"
            if (len(finalSolMoves[i]) == 2):
                if (finalSolMoves[i][1] == "\'"):
                    urlAlgCubing += finalSolMoves[i][0] + "-"
                else:
                    urlAlgCubing += finalSolMoves[i][0] + "2"
            else:
                urlAlgCubing += finalSolMoves[i]

        else:
            urlAlgCubing += "%0A%2F%2F_mistake_starts_here%2F%2F%0A"
            if (len(finalSolMoves[i]) == 2):
                if (finalSolMoves[i][1] == "\'"):
                    urlAlgCubing += finalSolMoves[i][0] + "-"
                else:
                    urlAlgCubing += finalSolMoves[i][0] + "2"
            else:
                urlAlgCubing += finalSolMoves[i]
        urlAlgCubing += "_"

    return  urlAlgCubing

def resetCube(Cube):
    Cube.solving = False
    Cube.moves = []
    Cube.solutionEdges = []
    Cube.solutionCorners = []
    Cube.moveNumber = 0
    Cube.alg.reset()
    Cube.moveNumber = 0
    Cube.scrambleToExe = getNewScramble()
    Cube.scrambleApplied = ""
    Cube.Recon = ""
    Cube.firstMove = False
    Cube.firstSolveMove = True
    Cube.scramble = True
    Cube.memo = False
    Cube.solving = False
    Cube.startMemoTime = time.time()
    Cube.memoTime = time.time()
    Cube.startExeTime = time.time()
    Cube.exeTime = time.time()
    Cube.isSolvedEdges = False
    Cube.isSolvedCorners = False
    Cube.bestArrayCor = []
    Cube.bestArrayEdge = []
    Cube.bestCorOrder = 0
    Cube.bestEdgeOrder = 0
    Cube.success = False
    Cube.rawSol = []
    Cube.mostPiecesSolved = 0





def mainFunc():

    #TODO: deal with parity
    #TODO: check following solves: success' mistakr in edges, mistake in corners,
    #TODO: deal with mistake with better success
    #TODO: EXCEL HYPERLYNC

    Cube = DNFanalyzer()
    resetCube(Cube)
    Cube.firstMove = True
    trainingTime = 300000.0



