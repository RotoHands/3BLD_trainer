#-*- coding: utf-8 -*-
import openpyxl
from openpyxl import load_workbook
wb = load_workbook('ROTO 3bld Algs.xlsx', data_only=True)
corAlgs = wb["corners"]
edAlgs = wb["edges"]
corAll = wb["allCorList"]
edAll = wb["allEdList"]
lettersCorners = ['א', 'ב', 'ד', 'ה', 'ו', 'ז', 'ח', 'ט', 'כ', 'ל', 'נ', 'ס', 'ע', 'פ', 'צ', 'ק', 'ר', 'ש', 'ת',
                  'צ\'', 'ג\'']
lettersEdges = ['א', 'ב', 'ד', 'ה', 'ו', 'ז', 'ח', 'י', 'כ', 'ל', 'מ', 'נ', 'ס', 'ע', 'פ', 'צ', 'ק', 'ר', 'ש', 'ת',
                'צ\'', 'ג\'']

count = 1
for i in range (2,24):
    for j in range(2, 24):
        if(edAlgs.cell(j,i).value != None):
            edAll.cell(count,1).value = lettersEdges[i-2] + lettersEdges[j-2]
            edAll.cell(count,2).value = edAlgs.cell(j,i).value
            count+=1

count = 1
for i in range (2,23):
    for j in range(2, 23):
        if(corAlgs.cell(j,i).value != None):
            corAll.cell(count,1).value = lettersCorners[i-2] + lettersCorners[j-2]
            corAll.cell(count,2).value = corAlgs.cell(j,i).value
            count+=1

wb.save('ROTO 3bld Algs.xlsx')