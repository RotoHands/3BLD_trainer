import websockets
import openpyxl
import pickle
class alg_detailed:
    def __init__(self, alg_string, letter_pair, piece, alg_index):
        self.alg_index = alg_index
        self.alg_string = alg_string
        self.letter_pair = letter_pair
        self.piece = piece
        self.solves_times = []
        self.recognize_times = []
        self.solves_times_len = None
        self.solves_times_avg = None
        self.recognize_times_len = None
        self.recognize_times_avg = None
        self.row_excel = None
        self.col_excel = None
        self.sheet_excel = None
        self.train_alg = False

    def __init__(self, alg_string, letter_pair, piece, row_excel, col_excel, sheet_excel, alg_index ):
        self.alg_index = alg_index
        self.alg_string = alg_string
        self.letter_pair = letter_pair
        self.piece = piece
        self.solves_times = []
        self.recognize_times = []
        self.row_excel = row_excel
        self.col_excel = col_excel
        self.sheet_excel = sheet_excel
        self.solves_times_len = None
        self.solves_times_avg = None
        self.recognize_times_len = None
        self.recognize_times_avg = None
        self.train_alg = True

    def __str__(self):
        return ("{} : {}\n{} : {}\n{} : {}\n{} : {}\n{} : {}\n{} : {}\n\n".format("alg_index", self.alg_index,"row", self.row_excel,"col", self.col_excel ,"letter_pair", self.letter_pair,"alg_string", self.alg_string, "piece", self.piece))
    def parse_alg_moves(self):
        pass


def init_algs_dict():
    wb = openpyxl.load_workbook('ROTO 3bld Algs.xlsx', data_only=True)
    edAlgs = wb["edges"]
    corAlgs = wb["corners"]
    lettersCorners = ['א', 'ב', 'ד', 'ה', 'ו', 'ז', 'ח', 'ט', 'כ', 'ל', 'נ', 'ס', 'ע', 'פ', 'צ', 'ק', 'ר', 'ש', 'ת',
                      'צ\'', 'ג\'']
    lettersEdges = ['א', 'ב', 'ד', 'ה', 'ו', 'ז', 'ח', 'י', 'כ', 'ל', 'מ', 'נ', 'ס', 'ע', 'פ', 'צ', 'ק', 'ר', 'ש', 'ת',
                    'צ\'', 'ג\'']
    letters = ['א', 'ב', 'ג', 'ד', 'ה', 'ו', 'ז', 'ח', 'ט', 'י', 'כ', 'ל', 'מ', 'נ', 'ס', 'ע', 'פ', 'צ', 'ק', 'ר', 'ש',         'ת', '1', '2']
    count = -1
    algs_dict = {}
    lp_2_index_edges = {}
    lp_2_corners_dict = {}

    for i in range(2,len(lettersEdges) +2):
        for j in range(2, len(lettersEdges) + 2):
            alg_string = edAlgs.cell(i,j).value

            if (alg_string != None):
                count += 1
                letter_pair = lettersEdges[j-2] + lettersEdges[i-2]
                current_alg = alg_detailed(alg_string, letter_pair, "e",i,j,"edges - drill", count)
                algs_dict[count] = current_alg
                lp_2_index_edges[letter_pair] = count

    for i in range(2,len(lettersCorners) +2):
        for j in range(2, len(lettersCorners) + 2):
            alg_string = corAlgs.cell(i,j).value
            if (alg_string != None):
                count += 1
                letter_pair = lettersCorners[j-2] + lettersCorners[i-2]
                current_alg = alg_detailed(alg_string, letter_pair, "c",i,j,"corners - drill", count)
                algs_dict[count] = current_alg
                lp_2_corners_dict[letter_pair] = count


    return algs_dict,lp_2_index_edges,lp_2_corners_dict

def load_pkl(path):
    with open(path, "rb") as f:
        return pickle.load(f)

def write_dict_to_pkl():
    with open("dicts.pkl", "wb") as f:
        dicts = init_algs_dict()
        pickle.dump(dicts, f)

def load_algs_dict(path):
    with open(path , "rb") as f:
        algs_dict,lp_2_index_edges,lp_2_corners_dict = pickle.load(f)

    return algs_dict,lp_2_index_edges,lp_2_corners_dict